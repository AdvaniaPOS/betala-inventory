"""
Excel-generatorer for rapporter.
"""

from io import BytesIO
from datetime import date, datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


class BaseExcelGenerator:
    """Baseklasse for Excel-generering."""
    
    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        
        # Stiler
        self.header_font = Font(bold=True, size=12)
        self.header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        self.header_font_white = Font(bold=True, size=12, color='FFFFFF')
        self.number_format = '#,##0'
        self.currency_format = '#,##0.00 "kr"'
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def set_header(self, row, values):
        """Sett header-rad med styling."""
        for col, value in enumerate(values, 1):
            cell = self.ws.cell(row=row, column=col, value=value)
            cell.font = self.header_font_white
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.thin_border
    
    def auto_width(self):
        """Auto-juster kolonnebredder."""
        for column in self.ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            self.ws.column_dimensions[column_letter].width = adjusted_width
    
    def save_to_buffer(self) -> BytesIO:
        """Lagre til BytesIO buffer."""
        output = BytesIO()
        self.wb.save(output)
        output.seek(0)
        return output


class InventoryExcelGenerator(BaseExcelGenerator):
    """Generator for lagerbeholdning-rapport."""
    
    def generate(self, event, stock_levels) -> BytesIO:
        self.ws.title = 'Lagerbeholdning'
        
        # Tittel
        self.ws['A1'] = f'Lagerbeholdning - {event.name}'
        self.ws['A1'].font = Font(bold=True, size=16)
        self.ws['A2'] = f'Generert: {datetime.now().strftime("%Y-%m-%d %H:%M")}'
        
        # Headers
        headers = ['Produkt', 'Kategori', 'Enhet', 'Beholdning', 'Min. nivå', 
                   'Status', 'Pris', 'Verdi']
        self.set_header(4, headers)
        
        # Data
        row = 5
        total_value = 0
        
        for sl in stock_levels:
            product = sl.product
            status = 'Lav' if sl.is_low_stock else 'OK'
            value = (product.price_ore or 0) * sl.quantity / 100
            total_value += value
            
            self.ws.cell(row=row, column=1, value=product.name)
            self.ws.cell(row=row, column=2, value=product.category.name if product.category else '')
            self.ws.cell(row=row, column=3, value=product.unit)
            self.ws.cell(row=row, column=4, value=sl.quantity)
            self.ws.cell(row=row, column=5, value=product.min_stock_level)
            
            status_cell = self.ws.cell(row=row, column=6, value=status)
            if status == 'Lav':
                status_cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            
            self.ws.cell(row=row, column=7, value=(product.price_ore or 0) / 100).number_format = self.currency_format
            self.ws.cell(row=row, column=8, value=value).number_format = self.currency_format
            
            row += 1
        
        # Sum
        row += 1
        self.ws.cell(row=row, column=7, value='Total:').font = Font(bold=True)
        self.ws.cell(row=row, column=8, value=total_value).number_format = self.currency_format
        self.ws.cell(row=row, column=8).font = Font(bold=True)
        
        self.auto_width()
        return self.save_to_buffer()


class TransactionExcelGenerator(BaseExcelGenerator):
    """Generator for transaksjonsrapport."""
    
    def generate(self, event, transactions, start_date) -> BytesIO:
        self.ws.title = 'Transaksjoner'
        
        # Tittel
        self.ws['A1'] = f'Transaksjoner - {event.name}'
        self.ws['A1'].font = Font(bold=True, size=16)
        self.ws['A2'] = f'Periode: {start_date} - {date.today()}'
        self.ws['A3'] = f'Generert: {datetime.now().strftime("%Y-%m-%d %H:%M")}'
        
        # Headers
        headers = ['Dato', 'Tid', 'Type', 'Produkt', 'Antall', 'Referanse', 'Bruker']
        self.set_header(5, headers)
        
        # Data
        row = 6
        for tx in transactions:
            self.ws.cell(row=row, column=1, value=tx.transaction_date.strftime('%Y-%m-%d'))
            self.ws.cell(row=row, column=2, value=tx.transaction_date.strftime('%H:%M'))
            self.ws.cell(row=row, column=3, value=tx.get_transaction_type_display())
            self.ws.cell(row=row, column=4, value=tx.product.name)
            
            qty_cell = self.ws.cell(row=row, column=5, value=tx.quantity)
            if tx.quantity < 0:
                qty_cell.font = Font(color='FF0000')
            else:
                qty_cell.font = Font(color='008000')
            
            self.ws.cell(row=row, column=6, value=tx.reference)
            self.ws.cell(row=row, column=7, value=tx.created_by.get_full_name() if tx.created_by else '')
            
            row += 1
        
        self.auto_width()
        return self.save_to_buffer()


class ShrinkageExcelGenerator(BaseExcelGenerator):
    """Generator for svinnrapport."""
    
    def generate(self, event, entries, start_date) -> BytesIO:
        self.ws.title = 'Svinn'
        
        # Tittel
        self.ws['A1'] = f'Svinnrapport - {event.name}'
        self.ws['A1'].font = Font(bold=True, size=16)
        self.ws['A2'] = f'Periode: {start_date} - {date.today()}'
        
        # Headers
        headers = ['Dato', 'Produkt', 'Antall', 'Årsak', 'Sted', 'Estimert tap', 'Registrert av']
        self.set_header(4, headers)
        
        # Data
        row = 5
        total_loss = 0
        
        for entry in entries:
            loss = (entry.estimated_loss_ore or 0) / 100
            total_loss += loss
            
            self.ws.cell(row=row, column=1, value=entry.registered_date.strftime('%Y-%m-%d %H:%M'))
            self.ws.cell(row=row, column=2, value=entry.product.name)
            self.ws.cell(row=row, column=3, value=entry.quantity)
            self.ws.cell(row=row, column=4, value=entry.get_reason_display())
            self.ws.cell(row=row, column=5, value=entry.location)
            self.ws.cell(row=row, column=6, value=loss).number_format = self.currency_format
            self.ws.cell(row=row, column=7, value=entry.registered_by.get_full_name() if entry.registered_by else '')
            
            row += 1
        
        # Sum
        row += 1
        self.ws.cell(row=row, column=5, value='Totalt tap:').font = Font(bold=True)
        self.ws.cell(row=row, column=6, value=total_loss).number_format = self.currency_format
        self.ws.cell(row=row, column=6).font = Font(bold=True)
        
        self.auto_width()
        return self.save_to_buffer()


# PDF-generatorer
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm, cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer


class StockCountPDFGenerator:
    """Generator for varetelling-rapport i PDF-format."""
    
    def __init__(self):
        self.styles = getSampleStyleSheet()
        self.title_style = ParagraphStyle(
            'CustomTitle',
            parent=self.styles['Heading1'],
            fontSize=16,
            spaceAfter=12,
        )
        self.subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=self.styles['Heading2'],
            fontSize=12,
            spaceAfter=6,
        )
        self.normal_style = ParagraphStyle(
            'CustomNormal',
            parent=self.styles['Normal'],
            fontSize=10,
        )
    
    def generate(self, stock_count) -> BytesIO:
        """Generer PDF for en varetelling, inkludert importerte deltellinger."""
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=1.5*cm,
            leftMargin=1.5*cm,
            topMargin=1.5*cm,
            bottomMargin=1.5*cm
        )
        
        elements = []
        
        # Tittel
        elements.append(Paragraph(f"Tellerapport: {stock_count.name}", self.title_style))
        
        # Metadata
        meta_data = [
            ['Event:', stock_count.event.name],
            ['Lokasjon:', stock_count.location or '-'],
            ['Startet:', stock_count.started_at.strftime('%d.%m.%Y %H:%M')],
            ['Startet av:', stock_count.started_by.get_full_name() if stock_count.started_by else '-'],
        ]
        if stock_count.completed_at:
            meta_data.append(['Fullført:', stock_count.completed_at.strftime('%d.%m.%Y %H:%M')])
            if stock_count.completed_by:
                meta_data.append(['Fullført av:', stock_count.completed_by.get_full_name()])
        
        meta_table = Table(meta_data, colWidths=[3*cm, 10*cm])
        meta_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(meta_table)
        elements.append(Spacer(1, 0.5*cm))
        
        # Importerte deltellinger (hvis noen)
        partial_counts = list(stock_count.partial_counts.filter(
            status='IMPORTED'
        ).order_by('completed_at'))
        
        if partial_counts:
            elements.append(Paragraph("Importerte deltellinger:", self.subtitle_style))
            partial_data = [['Navn', 'Lokasjon', 'Talt av', 'Tidspunkt']]
            for pc in partial_counts:
                teller = pc.completed_by.get_full_name() if pc.completed_by else (
                    pc.started_by.get_full_name() if pc.started_by else '-'
                )
                partial_data.append([
                    pc.name,
                    pc.location or '-',
                    teller,
                    pc.completed_at.strftime('%d.%m.%Y %H:%M') if pc.completed_at else '-'
                ])
            
            partial_table = Table(partial_data, colWidths=[5*cm, 4*cm, 4*cm, 4*cm])
            partial_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
            ]))
            elements.append(partial_table)
            elements.append(Spacer(1, 0.5*cm))
        
        # Hovedtelling produkter
        elements.append(Paragraph("Telleresultat:", self.subtitle_style))
        
        # Hent alle linjer - sortert etter kategori og produktnavn
        lines = stock_count.lines.select_related(
            'product', 'product__category'
        ).order_by('product__category__name', 'product__name')
        
        # Tabell med resultater
        table_data = [['Produkt', 'Kategori', 'Forventet', 'Talt', 'Avvik']]
        
        total_expected = 0
        total_counted = 0
        total_variance = 0
        
        for line in lines:
            expected = line.expected_quantity or 0
            counted = line.counted_quantity if line.counted_quantity is not None else '-'
            
            if line.counted_quantity is not None:
                variance = line.counted_quantity - expected
                total_expected += expected
                total_counted += line.counted_quantity
                total_variance += variance
                variance_str = f"+{variance}" if variance > 0 else str(variance)
            else:
                variance_str = '-'
            
            category = line.product.category.name if line.product.category else '-'
            
            table_data.append([
                line.product.name,
                category,
                str(expected),
                str(counted),
                variance_str
            ])
        
        # Sum-rad
        table_data.append([
            'TOTALT',
            '',
            str(total_expected),
            str(total_counted),
            f"+{total_variance}" if total_variance > 0 else str(total_variance)
        ])
        
        # Kolonnebredder
        col_widths = [6*cm, 4*cm, 2.5*cm, 2.5*cm, 2.5*cm]
        product_table = Table(table_data, colWidths=col_widths)
        
        # Styling
        style = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ALIGN', (2, 0), (-1, -1), 'CENTER'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            # Sum-rad styling
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#D9E2F3')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ]
        
        # Farg negative avvik rødt, positive grønne
        for i, row in enumerate(table_data[1:-1], start=1):
            variance_str = row[4]
            if variance_str != '-':
                try:
                    variance_val = int(variance_str.replace('+', ''))
                    if variance_val < 0:
                        style.append(('TEXTCOLOR', (4, i), (4, i), colors.red))
                    elif variance_val > 0:
                        style.append(('TEXTCOLOR', (4, i), (4, i), colors.green))
                except ValueError:
                    pass
        
        product_table.setStyle(TableStyle(style))
        elements.append(product_table)
        
        # Generer PDF
        doc.build(elements)
        buffer.seek(0)
        return buffer
    
    def generate_by_partial(self, stock_count) -> BytesIO:
        """Generer PDF sortert per deltelling."""
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=1.5*cm,
            leftMargin=1.5*cm,
            topMargin=1.5*cm,
            bottomMargin=1.5*cm
        )
        
        elements = []
        
        # Tittel
        elements.append(Paragraph(f"Tellerapport: {stock_count.name}", self.title_style))
        elements.append(Paragraph("(Sortert per deltelling)", self.normal_style))
        elements.append(Spacer(1, 0.3*cm))
        
        # Metadata
        meta_data = [
            ['Event:', stock_count.event.name],
            ['Lokasjon:', stock_count.location or '-'],
            ['Startet:', stock_count.started_at.strftime('%d.%m.%Y %H:%M')],
        ]
        if stock_count.completed_at:
            meta_data.append(['Fullført:', stock_count.completed_at.strftime('%d.%m.%Y %H:%M')])
        
        meta_table = Table(meta_data, colWidths=[3*cm, 10*cm])
        meta_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(meta_table)
        elements.append(Spacer(1, 0.5*cm))
        
        # Importerte deltellinger
        partial_counts = list(stock_count.partial_counts.filter(
            status='IMPORTED'
        ).order_by('completed_at'))
        
        col_widths = [6*cm, 4*cm, 2.5*cm, 2.5*cm, 2.5*cm]
        
        # Vis hver deltelling separat
        for idx, pc in enumerate(partial_counts, 1):
            teller = pc.completed_by.get_full_name() if pc.completed_by else (
                pc.started_by.get_full_name() if pc.started_by else '-'
            )
            
            elements.append(Paragraph(
                f"Deltelling {idx}: {pc.name}",
                self.subtitle_style
            ))
            
            # Deltelling info
            pc_meta = [
                ['Lokasjon:', pc.location or '-'],
                ['Talt av:', teller],
                ['Tidspunkt:', pc.completed_at.strftime('%d.%m.%Y %H:%M') if pc.completed_at else '-'],
            ]
            pc_meta_table = Table(pc_meta, colWidths=[3*cm, 10*cm])
            pc_meta_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
            ]))
            elements.append(pc_meta_table)
            elements.append(Spacer(1, 0.2*cm))
            
            # Produkter fra denne deltelling (kun de som har counted_quantity)
            pc_lines = pc.lines.filter(
                counted_quantity__isnull=False
            ).select_related('product', 'product__category').order_by('product__name')
            
            if pc_lines.exists():
                table_data = [['Produkt', 'Kategori', 'Forventet', 'Talt', 'Avvik']]
                
                for line in pc_lines:
                    expected = line.expected_quantity or 0
                    counted = line.counted_quantity
                    variance = counted - expected
                    variance_str = f"+{variance}" if variance > 0 else str(variance)
                    category = line.product.category.name if line.product.category else '-'
                    
                    table_data.append([
                        line.product.name,
                        category,
                        str(expected),
                        str(counted),
                        variance_str
                    ])
                
                pc_table = Table(table_data, colWidths=col_widths)
                
                style = [
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#5B9BD5')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 8),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('ALIGN', (2, 0), (-1, -1), 'CENTER'),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                    ('TOPPADDING', (0, 0), (-1, -1), 4),
                ]
                
                # Farg avvik
                for i, row in enumerate(table_data[1:], start=1):
                    try:
                        variance_val = int(row[4].replace('+', ''))
                        if variance_val < 0:
                            style.append(('TEXTCOLOR', (4, i), (4, i), colors.red))
                        elif variance_val > 0:
                            style.append(('TEXTCOLOR', (4, i), (4, i), colors.green))
                    except ValueError:
                        pass
                
                pc_table.setStyle(TableStyle(style))
                elements.append(pc_table)
            else:
                elements.append(Paragraph("Ingen produkter talt", self.normal_style))
            
            elements.append(Spacer(1, 0.5*cm))
        
        # Hovedtellingens egne linjer (hvis noen)
        main_lines = stock_count.lines.filter(
            counted_quantity__isnull=False
        ).select_related('product', 'product__category').order_by('product__name')
        
        if main_lines.exists():
            elements.append(Paragraph("Hovedtellingens egne produkter:", self.subtitle_style))
            
            table_data = [['Produkt', 'Kategori', 'Forventet', 'Talt', 'Avvik']]
            
            for line in main_lines:
                expected = line.expected_quantity or 0
                counted = line.counted_quantity
                variance = counted - expected
                variance_str = f"+{variance}" if variance > 0 else str(variance)
                category = line.product.category.name if line.product.category else '-'
                
                table_data.append([
                    line.product.name,
                    category,
                    str(expected),
                    str(counted),
                    variance_str
                ])
            
            main_table = Table(table_data, colWidths=col_widths)
            main_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ALIGN', (2, 0), (-1, -1), 'CENTER'),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
            ]))
            elements.append(main_table)
        
        # Generer PDF
        doc.build(elements)
        buffer.seek(0)
        return buffer
